from __future__ import print_function
from django.shortcuts import render, HttpResponseRedirect, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth.views import login
from VTAT_TOOL.forms import *
from VTAT_TOOL.models import Repository, RepositoryData, Genre, TestLogs
import json
from django.core.files.storage import FileSystemStorage
import sys


reload(sys)  
sys.setdefaultencoding('utf8')
import subprocess
from django.shortcuts import render_to_response
from django.template.loader import render_to_string
import os
import re
from django.core.files import File
import logging
import datetime
import time
#import threading
import json
import csv
import collections
# if sys.version_info >= (3,0):
try:
    import queue as queue
except ImportError:
    import Queue as queue
import ast
#import pandas as pd
from openpyxl import load_workbook
from django.conf import settings
import copy
import lxml.etree as ET
# import queue


# Create your views here.

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

work_book = load_workbook(BASE_DIR+"/config.xlsx")
work_sheet = work_book.worksheets[0]
tests_sheet = []
repo_name_sheet = (work_sheet.cell(row=2, column=1).value).capitalize()
categ_name_sheet = (work_sheet.cell(row=2, column=2).value).capitalize()
for j in range(2, work_sheet.max_row+1):
    tests_sheet.append((work_sheet.cell(row=j, column=3).value).lower())


@login_required
@csrf_exempt
def index(request):
    print (BASE_DIR)

    user_data = Repository.objects.filter(user=request.user.id)
    # print tests_sheet

    if user_data:
        html = render_to_string("index.html", {"repository_data": user_data, "status": 1,
                                               "repo_name_sheet": repo_name_sheet, "categ_name_sheet": categ_name_sheet, "tests_sheet": tests_sheet})
        return HttpResponse(html)
        # return
        # render(request,"index.html",{"repository_data":user_data,"status":1})
    else:
        html_data = []
        html = render_to_string("index.html", {"repository_data": html_data, "status": 2,
                                               "repo_name_sheet": repo_name_sheet, "categ_name_sheet": categ_name_sheet, "tests_sheet": tests_sheet})
        return HttpResponse(html)
        # return
        # render(request,"index.html",{"repository_data":user_data,"status":1})


@login_required
@csrf_exempt
def create_repository(request):
    if request.method == "POST":
        html = render_to_string("addRepository.html", {
                                "repo_name_sheet": repo_name_sheet, "categ_name_sheet": categ_name_sheet, "tests_sheet": tests_sheet})
        return HttpResponse(json.dumps({"status": 1, "html_data": html}))
    else:
        return HttpResponse(json.dumps({"status": 0, "html_data": "/NotFound/"}))


def file_not_found_error(request):
    html = render_to_string("error.html", {})
    return HttpResponse(html)


@csrf_exempt
def add_repository(request):
    if request.method == "POST":
        # try:
        repository_name = request.POST.get("repository_name")
        repository_description = request.POST.get("repository_description")
        file_upload = request.FILES.getlist("mysourcefile")
        '''match = re.search(r'[!@#-~^.]+', repository_name)
        if match:
            return HttpResponse({"status": 1, "html_data": "Repository Name Shouldn't contain Special Characters"})
           '''
        # print(repository_name,repository_description,file_upload)
        r = Repository(user_id=request.user.id, repository_name=repository_name,
                       repository_description=repository_description)
        for filename in request.FILES.getlist("mysourcefile"):
            r.repository_file = filename
        for filename in request.FILES.getlist("myreadmefile"):
            r.repository_readme = filename
        r.save()

        repo_obj = Repository.objects.filter(id=r.id)
        if repo_obj:
            repo_obj = Repository.objects.get(id=r.id)
        else:
            return HttpResponse(json.dumps({"status": 1, "html_data": "Error in creating Repository,Please Try Again!"}))

        root = ET.parse(repo_obj.repository_file.path)

        for ele in root.iter():
            if ele.getparent() is not None:
                parent_name = ele.getparent().attrib.get("name")
                parent_id = RepositoryData.objects.filter(
                    testcase_name=parent_name, repository_id=r.id)
                if len(parent_id) == 1:
                    RepositoryData.objects.create(repository_id=r.id, parent=parent_id[0], testcase_name=ele.attrib.get("name"), test_iterations=ele.attrib.get(
                        "iterations", "1"), testcase_action=ele.attrib.get("action", " "), log_name=ele.attrib.get("logfile", " "))
                    pass
                else:
                    return HttpResponse(json.dumps({"status": 1, "html_data": "TestCase name Should be Unique,Please Check"}))
            else:
                # print(r.id)
                # print("name",ele.attrib.get("name"))
                # print("iter",ele.attrib.get("iterations","100"))
                # print("iter",ele.attrib.get("action"," "))

                RepositoryData.objects.create(repository_id=r.id, testcase_name=ele.attrib.get("name"), test_iterations=ele.attrib.get(
                    "iterations", "1"), testcase_action=ele.attrib.get("action", " "), log_name=ele.attrib.get("logfile", " "))

        return HttpResponse(json.dumps({"status": 1, "html_data": "Suucess"}))
        # except Exception:
        # return HttpResponse(json.dumps({"status":0,"html_data":"error"}))

    else:
        return HttpResponse(json.dumps({"status": 0, "html_data": "error"}))


@csrf_exempt
def view_repository(request):
    if request.method == "POST":
        repo_id = request.POST.get("repo_id")
        # print("repo_id",repo_id)
        request.session["repo_id"] = repo_id
        repo_data = RepositoryData.objects.filter(repository_id=repo_id)
        def serializable_object(node):
            "Recurse into tree to build a serializable object"
            obj = collections.OrderedDict(
                {'name': node.testcase_name, 'id': node.pk, 'children': []})
            for child in node.get_children():
                obj['children'].append(serializable_object(child))
            return obj
        nodes = serializable_object(
            RepositoryData.objects.get(id=repo_data[0].id))
        html = render_to_string("view_repository2.html", {"nodes_data": json.dumps(nodes), "repo_id": request.session[
                                "repo_id"], "repo_name_sheet": repo_name_sheet, "categ_name_sheet": categ_name_sheet, "tests_sheet": tests_sheet})

        #html = render_to_string("view_repository2.html",{"repo_data":repo_data,"repo_id":request.session["repo_id"],"module_list":json.dumps(module_list),"repo_name_sheet":repo_name_sheet,"categ_name_sheet":categ_name_sheet,"tests_sheet":tests_sheet})
        return HttpResponse(json.dumps({"status": 1, "html_data": html}))
        # return HttpResponse(html)
    else:
        return HttpResponse(json.dumps({"status": 0, "html_data": "/NotFound/"}))


@csrf_exempt
def execute_repository(request):
    if request.method == "POST":
        repo_id = request.session["repo_id"]
        nodes = RepositoryData.objects.filter(repository_id=repo_id)

        html = render_to_string("execute_repository.html", {
                                "nodes": nodes, "repo_name_sheet": repo_name_sheet, "categ_name_sheet": categ_name_sheet, "tests_sheet": tests_sheet})
        return HttpResponse(json.dumps({"status": 1, "html_data": html}))
    else:
        return HttpResponse(json.dumps({"status": 0, "html_data": "/NotFound/"}))

global q
q = queue.Queue()
global run_flag
global monitor_list
monitor_list = []
run_flag = 0


@csrf_exempt
def run_repository(request):
    if request.method == "POST":
        global run_flag
        global q
        global monitor_list
        module_array = json.loads(request.POST.get("module_array"))
        module_iterations = request.POST.get("module_iterations")
        # print(module_array,module_iterations)

        def run_script_multiple(file_id,log_fp):
            print("*************************")
            success_iterations = 0
            fail_iterations = 0
            file_ide = RepositoryData.objects.filter(id=file_id)

            if file_ide:
                print("11111111111111111111111111111")
                if file_ide:
                    file_obj = RepositoryData.objects.get(id=file_id)
                    cmd = file_obj.testcase_action
                    
                    if cmd == " ":
                        return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 1}
                    parent_names = []
                    def parent_name(parentID):
                        parentID = RepositoryData.objects.get(id=parentID)
                        if parentID.parent:
                            parent_names.append(parentID.parent.testcase_name)
                            return parent_name(parentID.parent_id)
                        else:
                            return  parent_names

                    parent_names = parent_name(file_obj.id)
                    print ("parent_names",parent_names)


                    #if file_obj.parent and file_obj.parent.testcase_name == "runtest":
                    if "runtest" in parent_names:
                        print ("ipcipcipcipcipicpipcpipcipicpicp")
                        os.chdir("/opt/ltp")
                        print (os.getcwd())
                        cmd = file_obj.testcase_action
                        print("cmd",cmd)
                    elif file_obj.parent and file_obj.parent.testcase_name == "testscripts":
                        os.chdir("/opt/ltp/testscripts")
                        cmd = file_obj.testcase_action
                    else:
                        pass
                    
		    #test_parent_name = parent_names[0]
                    #pro_dir = parent_names[-1]
                    parent_names = []
		    print("cmd", cmd)
                    #cur_dir = os.getcwd()
                    #os.chdir(cur_dir+'/'+pro_dir)
                    print(os.getcwd())
                    print("88888888888888888888888")
                    if os.path.isfile(file_obj.log_name):
                        os.remove(file_obj.log_name)
                    process = subprocess.Popen(
                        cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, shell=True)
                    out, err = process.communicate()

                    print ("out is",out)
                    print ("err is",err)
                    print(file_obj.log_name)
                    print("//////////////////////////////////")
                    
                    with open(str(file_obj.log_name), "a") as fr:
                        fr.write('\nTest Name: {}'.format(test_parent_name))
                        fr.write('\nSubTest Name: {}\n'.format(file_obj.testcase_name))
                        # fr.close()
                        
                    with open(str(file_obj.log_name), "r") as fr:
                        lines = fr.readlines()


                        #print("lines", lines)
                        if lines:
                            lines_data = ''.join(lines)
                            # lines_data = ''
                            # for line_data in lines:
                            #     lines_data=lines_data+line_data
                            log_fp.writelines(str(lines_data))#.encode('utf-8').strip())
                            if re.search(r"error|exception", fr.read(), re.I):
                                print("re")
                                fail_iterations += 1
                                return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 0}
                            else:
                                print("err,", err)
                                if str(err).encode('utf-8') != '':
                                    log_fp.write(str(err).encode('utf-8'))
                                    fail_iterations += 1
                                    return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 0}
                                else:
                                    success_iterations += 1
                                    return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 1}
                        else:
                            fail_iterations += 1
                            return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 0}
                        #fr.close()
                        os.chdir(BASE_DIR)
                # except Exception as e:
                    #fail_iterations +=1
                    # return
                    # {"success_iterations":success_iterations,"fail_iterations":fail_iterations,"status":-1}

            else:
                fail_iterations += 1
                return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": -1}

        query_set = []
        for ele in module_array:
            obj = RepositoryData.objects.filter(id=ele)
            if obj:

                if obj.get_ancestors():
                    for ele in obj.get_ancestors(include_self=True):
                        query_set.append(ele)
                if obj.get_descendants():
                    for ele in obj.get_descendants(include_self=True):
                        query_set.append(ele)

        # queue_list = list(set(query_set))
        queue_list = []
        for val in query_set:
            if val not in queue_list:
                queue_list.append(val)
        print(queue_list)

        for query_ele in queue_list:
            # print(query_ele,type(query_ele))
            q.put({"file_ids": query_ele, "iterations": module_iterations,
                   "repo_id": request.session["repo_id"], "status": "Scheduled"})

            monitor_list.append({"file_ids": query_ele, "iterations": module_iterations,
                                 "repo_id": request.session["repo_id"], "status": "Scheduled"})
        # for i in monitor_list:
        # 	print(i)
        # print(monitor_list)

        if run_flag == 0:
            while not q.empty():
                run_flag = 1
                dic = q.get()
                try:
                    dic_index = monitor_list.index(dic)
                except Exception as e:
                    continue
                if monitor_list[dic_index]["status"] == "Scheduled":
                    monitor_list[dic_index]["status"] = "Running"

                    # print(str(dic["file_ids"].id))
                    for index_iteration, i in enumerate(range(int(dic["iterations"]))):
                        plan_exec_start = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        log_file = "/tmp/file_" + \
                            str(dic["file_ids"].id)+"_"+plan_exec_start+".txt"
                        log_fp = open(log_file, "a+")

                        script_res = run_script_multiple(dic["file_ids"].id, log_fp)
                        print("script result", script_res)
                        plan_exec_end = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        log_fp.close()
                        if script_res["status"] == 1:
                            monitor_list[dic_index]["status"] = "Completed"
                            fp = open(log_file)
                            result = TestLogs()
                            result.log_testcase_id = dic["file_ids"].id
                            result.log_file_name.save(
                                "file_"+str(dic["file_ids"].id)+"_"+plan_exec_start+".txt", File(fp))
                            fp.close()
                            result.success_iterations = script_res[
                                "success_iterations"]
                            result.Fail_iterations = script_res[
                                "fail_iterations"]
                            result.script_exec_start = plan_exec_start
                            result.script_exec_end = plan_exec_end
                            result.script_status = "Pass"
                            result.save()
                            continue
                        else:
                            monitor_list[dic_index]["status"] = "Failed"
                            result = TestLogs()
                            fp = open(log_file)
                            result.log_testcase_id = dic["file_ids"].id
                            result.log_file_name.save(
                                "file_"+str(dic["file_ids"].id)+"_"+plan_exec_start+".txt", File(fp))
                            fp.close()
                            result.success_iterations = script_res[
                                "success_iterations"]
                            result.Fail_iterations = script_res[
                                "fail_iterations"]
                            result.script_exec_start = plan_exec_start
                            result.script_exec_end = plan_exec_end
                            result.script_status = "Fail"
                            result.save()

                            file_id = RepositoryData.objects.filter(
                                id=dic["file_ids"].id)
                            if file_id.get_descendants():
                                file_list = file_id.get_descendants()
                                for ide in file_list:
                                    for ele in monitor_list:
                                        if ele["file_ids"] == ide:
                                            ele_index = monitor_list.index(ele)
                                            monitor_list[ele_index][
                                                "status"] = "Parent "+dic["file_ids"].testcase_name+"Failed"
                                            result = TestLogs()
                                            result.log_testcase_id = ele[
                                                "file_ids"].id
                                            result.success_iterations = 0
                                            result.Fail_iterations = 0
                                            result.script_exec_start = ''
                                            result.script_exec_end = ''
                                            result.script_status = "Parent " + \
                                                dic["file_ids"].testcase_name + \
                                                "Failed"
                                            result.save()

                else:
                    continue
            run_flag = 0

        return HttpResponse("success")
    else:
        return HttpResponse("error")


@csrf_exempt
def monitor_execution(request):
    if request.method == "POST":
        global monitor_list
        #monitor_list = []
        # monitor_list.append({"file_ids":"testcase","iterations":5,"repo_id":request.session["repo_id"],"status":"Scheduled"})

        show_monitor_list = copy.deepcopy(monitor_list)
        html = render_to_string("monitor_execution.html", {
                                "monitor_list": monitor_list,"repo_name_sheet": repo_name_sheet, "categ_name_sheet": categ_name_sheet, "tests_sheet": tests_sheet})
        return HttpResponse(json.dumps({"status": 1, "html_data": html}))
    else:
        return HttpResponse(json.dumps({"status": 0, "html_data": "/NotFound/"}))
#from mptt.forms import TreeNodeChoiceField


@csrf_exempt
def results_repository(request):
    if request.method == "POST":
        repo_id = request.session["repo_id"]
        nodes = RepositoryData.objects.filter(repository_id=repo_id)
        #nodes = TreeNodeChoiceField(queryset=RepositoryData.objects.filter(repository_id=repo_id),level_indicator=u'+--')
        nodes_dict = collections.OrderedDict()
        for node in nodes:
            obj = TestLogs.objects.filter(log_testcase_id=node.id)
            if obj:
                nodes_dict[node] = TestLogs.objects.filter(
                    log_testcase_id=node.id).latest("id")
            else:
                nodes_dict[node] = []

        print(nodes_dict)

        html = render_to_string("results_repository.html", {
                                "nodes": nodes_dict, "repo_name_sheet": repo_name_sheet, "categ_name_sheet": categ_name_sheet, "tests_sheet": tests_sheet})
        return HttpResponse(json.dumps({"status": 1, "html_data": html}))
    else:
        return HttpResponse(json.dumps({"status": 0, "html_data": "/NotFound/"}))


@csrf_exempt
def show_script_logs(request):
    if request.method == "POST":
        fileid = request.POST.get("file_id")
        file = TestLogs.objects.filter(id=fileid)
        if file:
            file_obj = TestLogs.objects.get(id=fileid)
            if file_obj.log_file_name:
                res_data = file_obj.log_file_name.chunks()
                data = ''
                for i in res_data:
                    print(type(i),i[0])
                    try:
                        data = data+str(i).encode('utf-8')

                    except UnicodeDecodeError:
                        data = data+str(i)

                data = data.replace("\n", '<br>')
                
                if len(data) == 0:
                    return HttpResponse("No information Provided")
                else:
                    return HttpResponse(data)
            else:
                return HttpResponse("NO Data Found")
        else:
            return HttpResponse("NO Data Found")
    else:
        return HttpResponse("NO Data Found")


@csrf_exempt
def show_monitor_script_logs(request):
    if request.method == "POST":
        fileid = request.POST.get("file_id")
        test_obj = RepositoryData.objects.filter(id=fileid)
        if test_obj:
            file_obj = TestLogs.objects.filter(log_testcase_id=fileid)
            if file_obj:
                file_obj = TestLogs.objects.filter(
                    log_testcase_id=fileid).latest("id")
                if file_obj.log_file_name:
                    res_data = file_obj.log_file_name.chunks()
                    data = ''
                    for i in res_data:
                        try:
                            data = data+str(i).encode('utf-8')
                        except UnicodeDecodeError:
                            data = data+str(i)
                    data = data.replace("\n", '<br>')
                    if len(data) == 0:
                        return HttpResponse("No information Provided")
                    else:
                        return HttpResponse(data)
                else:
                    return HttpResponse("NO Data Found")
            else:
                return HttpResponse("NO Data Found")
        else:
            return HttpResponse("NO Data Found")
    else:
        return HttpResponse("")


@csrf_exempt
def delete_repository(request):
    if request.method == "POST":
        repo_id = request.POST.get("repo_id")
        repo = Repository.objects.filter(id=repo_id)
        if repo:
            nodes = RepositoryData.objects.filter(repository_id=repo_id)
            if nodes:
                RepositoryData.objects.filter(repository_id=repo_id).delete()
            Repository.objects.filter(id=repo_id).delete()
        else:
            return HttpResponse("success")
    else:
        return HttpResponse("Try Again")

@csrf_exempt
def modify_repository(request):
	if request.method == "POST":
		repo_id = request.POST.get("repo_id")
		print("repo_id",repo_id)
		request.session["repo_id"] = repo_id
		#repo_data = RepositoryModules.objects.filter(repository=repo_id)
		
		html = render_to_string("modify_repository.html",{"repo_name_sheet":repo_name_sheet,"categ_name_sheet":categ_name_sheet,"tests_sheet":tests_sheet})
		return HttpResponse(json.dumps({"status":1,"html_data":html}))
	else:
		return HttpResponse(json.dumps({"status":0,"html_data":"/NotFound/"}))


@csrf_exempt
def monitor_clear_records(request):
    if request.method == "POST":
        global monitor_list

        for ele in monitor_list:
            if ele["status"] == "Scheduled":
                return HttpResponse(json.dumps({"status":2,"html_data":"Still the Execution is going,we can't clear records"}))
        
        monitor_list = []
        html = render_to_string("monitor_execution.html", {"monitor_list": monitor_list,"repo_name_sheet": repo_name_sheet, "categ_name_sheet": categ_name_sheet, "tests_sheet": tests_sheet})

        return HttpResponse(json.dumps({"status":1,"html_data":html}))
    else:
        return HttpResponse(json.dumps({"status":0,"html_data":"html"}))



@csrf_exempt
def DeviceBranchList(request):
    """if 1:
        # def get(self, request, *args, **kwargs):
        repo_id = request.session["repo_id"]
        branches = RepositoryData.objects.filter(
            repository_id=repo_id).exclude(mptt_level=0).order_by('lft')
        mptt_level_max = RepositoryData.objects.aggregate(Max('mptt_level'))
        i_max = mptt_level_max['mptt_level__max']
        test = transMPTTTable(RepositoryData, branches, i_max)
        return render(request, 'devicebranch_list.html', {'test': test})

    """
    import os
    import re
    """os.chdir("/home/naluvjyo/Documents/git_train/project")
    7 cmd = "git diff --name-only"
    8 p=Popen(cmd, stdout=PIPE, stderr=PIPE, shell=True)
    9 out,err = p.coomunicate()"""
    import git

    def run_script_multiple(file_id,log_fp):
            success_iterations = 0
            fail_iterations = 0
            file_ide = RepositoryData.objects.filter(id=file_id)

            if file_ide:
                if file_ide:
                    file_obj = RepositoryData.objects.get(id=file_id)
                    cmd = file_obj.testcase_action
                    
                    if cmd == " ":
                        return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 1}
                    parent_names = []
                    def parent_name(parentID):
                        parentID = RepositoryData.objects.get(id=parentID)
                        if parentID.parent:
                            parent_names.append(parentID.parent.testcase_name)
                            return parent_name(parentID.parent_id)
                        else:
                            return  parent_names

                    parent_names = parent_name(file_obj.id)
                    print ("parent_names",parent_names)


                    #if file_obj.parent and file_obj.parent.testcase_name == "runtest":
                    if "runtest" in parent_names:
                        os.chdir("/opt/ltp")
                        print (os.getcwd())
                        cmd = file_obj.testcase_action
                        print("cmd",cmd)
                    elif file_obj.parent and file_obj.parent.testcase_name == "testscripts":
                        os.chdir("/opt/ltp/testscripts")
                        cmd = file_obj.testcase_action
                    else:
                        pass
                    test_parent_name = parent_names[0]
                    parent_names = []


                    print("cmd", cmd)
                    print("88888888888888888888888")
                    if os.path.isfile(file_obj.log_name):
                        os.remove(file_obj.log_name)
                    process = subprocess.Popen(
                        cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, shell=True)
                    out, err = process.communicate()

                    print ("out is",out)
                    print ("err is",err)
                    print(file_obj.log_name)
                    print("//////////////////////////////////")
                    
                    with open(str(file_obj.log_name), "a") as fr:
                        fr.write('\nTest Name: {}'.format(test_parent_name))
                        fr.write('\nSubTest Name: {}\n'.format(file_obj.testcase_name))
                        fr.close()
                    
                    with open(str(file_obj.log_name), "r") as fr:
                        lines = fr.readlines()


                        #print("lines", lines)
                        if lines:
                            lines_data = ''.join(lines)
                            # lines_data = ''
                            # for line_data in lines:
                            #     lines_data=lines_data+line_data
                            log_fp.writelines(str(lines_data))#.encode('utf-8').strip())
                            if re.search(r"error|exception", fr.read(), re.I):
                                print("re")
                                fail_iterations += 1
                                return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 0}
                            else:
                                print("err,", err)
                                if str(err).encode('utf-8') != '':
                                    log_fp.write(str(err).encode('utf-8'))
                                    fail_iterations += 1
                                    return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 0}
                                else:
                                    success_iterations += 1
                                    return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 1}
                        else:
                            fail_iterations += 1
                            return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": 0}
                        #fr.close()
                        os.chdir(BASE_DIR)
                # except Exception as e:
                    #fail_iterations +=1
                    # return
                    # {"success_iterations":success_iterations,"fail_iterations":fail_iterations,"status":-1}

            else:
                fail_iterations += 1
                return {"success_iterations": success_iterations, "fail_iterations": fail_iterations, "status": -1}



    repo=git.Repo("/home/naluvjyo/Documents/git_train/project")
    hcommit = repo.head.commit
    print(hcommit)
    print(hcommit.diff('HEAD~1'))
    #print(dir(hcommit.diff('HEAD~1')))
    #print(dir(hcommit))
    print(hcommit.parents[0])
    #print(dir(hcommit.parents))

    # x=repo.git.diff(hcommit,hcommit.parents[0])
    # print(dir(x))
    dic = {}
    for diff_added in hcommit.diff('HEAD~1').iter_change_type('M'):
        #print(diff_added)
        #print(dir(diff_added))
        print ("\n")
        #print (diff_added.a_path)
        x=repo.git.diff(hcommit,hcommit.parents[0],diff_added.a_path)
        print(str(x))
        #print('\n'.join(str(x).split("\n")))

        #print(dir(x))
        for i in str(x).split("\n"):
            match = re.match(r'^-[a-zA-Z0-9_\s]+',i,re.DOTALL)
            if match:
                testcase = diff_added.a_path.split("/")[-1]
                print (match)
                if dic.has_key(testcase):
                    dic[testcase].append(match.group()[1:])
                else:
                    dic[testcase]=[]
                    dic[testcase].append(match.group()[1:])


    print (dic.items())
    print ("success")
    print(request.session["repo_id"])
    repo = Repository.objects.get(id=request.session["repo_id"])
    #for key,val in dic.items():
    new_testcase_ids = []
    for key,value in dic.items():
        if key == "fs":
            repo_modules = RepositoryData.objects.get(repository_id=repo.id,testcase_name="filesystem")
        else:
            repo_modules = RepositoryData.objects.get(repository_id=repo.id,testcase_name=key)
        #print (repo_modules.get_children())
        #print (repo_modules.get_descendant_count())
        for testcase_val in value:
            log_file = "/tmp/"+testcase_val.replace(" ","_")+".log"
            test_action = "./runltp -p -l "+log_file+" -f "+key+" -s "+testcase_val
            print (test_action)
            print(log_file)
            r=RepositoryData.objects.create(repository_id=repo.id, parent=repo_modules, testcase_name=testcase_val, test_iterations="1", testcase_action=test_action, log_name=log_file)
            r.save()
            new_testcase_ids.append(r.id)

    for file_id in new_testcase_ids:
        plan_exec_start = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_file = "/tmp/file_" + str(file_id)+"_"+plan_exec_start+".txt"
        log_fp = open(log_file, "a+")
        script_res = run_script_multiple(file_id, log_fp)
        print("script result", script_res)
        plan_exec_end = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_fp.close()

        fp = open(log_file)
        result = TestLogs()
        result.log_testcase_id = file_id
        result.log_file_name.save("file_"+str(file_id)+"_"+plan_exec_start+".txt", File(fp))
        fp.close()
        result.success_iterations = script_res["success_iterations"]
        result.Fail_iterations = script_res["fail_iterations"]
        result.script_exec_start = plan_exec_start
        result.script_exec_end = plan_exec_end

        if script_res["status"] == 1:
            result.script_status = "Pass"
        else:
            result.script_status = "Pass"
        result.save()
        continue


    


    
    #return HttpResponse(repo_modules.id)
    return HttpResponse("html")
